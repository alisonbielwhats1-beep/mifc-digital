from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def color_value(color: Any) -> str | None:
    if color is None:
        return None
    kind = getattr(color, "type", None)
    if kind == "rgb":
        return getattr(color, "rgb", None)
    if kind == "indexed":
        return f"indexed:{getattr(color, 'indexed', None)}"
    if kind == "theme":
        tint = getattr(color, "tint", 0)
        return f"theme:{getattr(color, 'theme', None)}:tint:{tint}"
    return kind


def cell_style(cell: Any) -> dict[str, Any]:
    fill = cell.fill
    return {
        "style_id": cell.style_id,
        "fill_type": fill.fill_type,
        "fill_fg": color_value(fill.fgColor),
        "font_color": color_value(cell.font.color),
        "locked": cell.protection.locked,
        "number_format": cell.number_format,
    }


def validation_map(ws: Any) -> dict[str, list[str]]:
    result: dict[str, list[str]] = {}
    validations = getattr(ws.data_validations, "dataValidation", [])
    for validation in validations:
        descriptor = f"{validation.type}:{validation.formula1 or ''}:{validation.formula2 or ''}"
        for cell_range in validation.ranges.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                for cell in row:
                    result.setdefault(cell.coordinate, []).append(descriptor)
    return result


def sheet_summary(ws: Any) -> dict[str, Any]:
    formulas = 0
    constants = 0
    nonempty = 0
    unlocked_nonempty = 0
    unlocked_blank_styled = 0
    comments = 0
    hyperlinks = 0
    style_counts: Counter[str] = Counter()
    formula_samples: list[dict[str, Any]] = []
    unlocked_samples: list[dict[str, Any]] = []

    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value is not None:
                nonempty += 1
                if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
                    formulas += 1
                    if len(formula_samples) < 12:
                        formula_samples.append({"cell": cell.coordinate, "formula": value})
                    kind = "formula"
                else:
                    constants += 1
                    kind = "constant"
                if not cell.protection.locked:
                    unlocked_nonempty += 1
                    if len(unlocked_samples) < 20:
                        unlocked_samples.append({"cell": cell.coordinate, "value": value})
                style = cell_style(cell)
                style_counts[
                    json.dumps(
                        {
                            "kind": kind,
                            "style_id": style["style_id"],
                            "fill_type": style["fill_type"],
                            "fill_fg": style["fill_fg"],
                            "locked": style["locked"],
                            "number_format": style["number_format"],
                        },
                        ensure_ascii=False,
                        sort_keys=True,
                    )
                ] += 1
            elif cell.has_style and not cell.protection.locked:
                unlocked_blank_styled += 1
            if cell.comment is not None:
                comments += 1
            if cell.hyperlink is not None:
                hyperlinks += 1

    validations = getattr(ws.data_validations, "dataValidation", [])
    conditional_rules = sum(len(rules) for rules in ws.conditional_formatting._cf_rules.values())
    print_area = None
    if ws.print_area:
        print_area = str(ws.print_area)

    return {
        "title": ws.title,
        "state": ws.sheet_state,
        "dimension": ws.calculate_dimension(),
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "nonempty": nonempty,
        "constants": constants,
        "formulas": formulas,
        "unlocked_nonempty": unlocked_nonempty,
        "unlocked_blank_styled": unlocked_blank_styled,
        "merged_range_count": len(ws.merged_cells.ranges),
        "data_validation_count": len(validations),
        "conditional_rule_count": conditional_rules,
        "image_count": len(getattr(ws, "_images", [])),
        "chart_count": len(getattr(ws, "_charts", [])),
        "comment_count": comments,
        "hyperlink_count": hyperlinks,
        "print_area": print_area,
        "sheet_protection": bool(ws.protection.sheet),
        "hidden_rows": sum(1 for item in ws.row_dimensions.values() if item.hidden),
        "hidden_columns": sum(1 for item in ws.column_dimensions.values() if item.hidden),
        "formula_samples": formula_samples,
        "unlocked_samples": unlocked_samples,
        "top_style_profiles": [
            {"count": count, **json.loads(profile)}
            for profile, count in style_counts.most_common(15)
        ],
    }


def workbook_summary(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False, read_only=False, keep_links=True)
    calc = getattr(wb, "calculation", None)
    external_links = getattr(wb, "_external_links", [])
    result = {
        "path": str(path),
        "sheet_names": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
        "defined_names": sorted(name.name for name in wb.defined_names.values()),
        "external_link_count": len(external_links),
        "calculation": {
            "mode": getattr(calc, "calcMode", None),
            "full_calc_on_load": getattr(calc, "fullCalcOnLoad", None),
            "force_full_calc": getattr(calc, "forceFullCalc", None),
        },
        "sheets": [sheet_summary(ws) for ws in wb.worksheets],
    }
    wb.close()
    return result


def dump_sheet(
    path: Path,
    sheet_name: str,
    only_unlocked: bool,
    include_blank: bool,
    cell_range: str | None,
) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False, read_only=False, keep_links=True)
    values_wb = load_workbook(path, data_only=True, read_only=False, keep_links=False)
    ws = wb[sheet_name]
    values_ws = values_wb[sheet_name]
    validations = validation_map(ws)
    cells = []
    if cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        rows = ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
    else:
        rows = ws.iter_rows()
    for row in rows:
        for cell in row:
            if only_unlocked and cell.protection.locked:
                continue
            if cell.value is None and not (include_blank and cell.has_style):
                continue
            record = {
                "cell": cell.coordinate,
                "row": cell.row,
                "column": cell.column,
                "value": cell.value,
                "cached_value": values_ws[cell.coordinate].value,
                "data_type": cell.data_type,
                "validation": validations.get(cell.coordinate, []),
                **cell_style(cell),
            }
            cells.append(record)
    result = {
        "path": str(path),
        "sheet": sheet_name,
        "dimension": ws.calculate_dimension(),
        "cells": cells,
    }
    wb.close()
    values_wb.close()
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbooks", nargs="+", type=Path)
    parser.add_argument("--sheet")
    parser.add_argument("--only-unlocked", action="store_true")
    parser.add_argument("--include-blank", action="store_true")
    parser.add_argument("--compact", action="store_true")
    parser.add_argument("--range")
    parser.add_argument("--tsv", action="store_true")
    args = parser.parse_args()

    if args.sheet:
        if len(args.workbooks) != 1:
            raise SystemExit("--sheet accepts exactly one workbook")
        output = dump_sheet(
            args.workbooks[0],
            args.sheet,
            args.only_unlocked,
            args.include_blank,
            args.range,
        )
    else:
        output = [workbook_summary(path) for path in args.workbooks]
        if args.compact:
            compact_keys = (
                "title",
                "state",
                "dimension",
                "nonempty",
                "constants",
                "formulas",
                "unlocked_nonempty",
                "unlocked_blank_styled",
                "data_validation_count",
                "conditional_rule_count",
                "image_count",
                "chart_count",
                "sheet_protection",
                "hidden_rows",
                "hidden_columns",
                "print_area",
            )
            output = [
                {
                    "path": workbook["path"],
                    "sheet_count": workbook["sheet_count"],
                    "external_link_count": workbook["external_link_count"],
                    "calculation": workbook["calculation"],
                    "sheets": [
                        {key: sheet[key] for key in compact_keys}
                        for sheet in workbook["sheets"]
                    ],
                }
                for workbook in output
            ]
    if args.tsv and args.sheet:
        print("cell\tvalue\tcached\ttype\tstyle\tfill\tlocked\tformat\tvalidation")
        for cell in output["cells"]:
            values = (
                cell["cell"],
                cell["value"],
                cell["cached_value"],
                cell["data_type"],
                cell["style_id"],
                cell["fill_fg"],
                cell["locked"],
                cell["number_format"],
                " | ".join(cell["validation"]),
            )
            print("\t".join(str(value).replace("\t", " ").replace("\n", " ") for value in values))
    else:
        print(json.dumps(output, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
